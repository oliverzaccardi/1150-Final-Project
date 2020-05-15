"""1150 Final Project. This program uses the beautiful soup package to parse through the MCTC Itec course search results
web page and adds all of the class information to an excel spreadsheet."""
# imports requests library
import requests
# imports beautiful soup package
from bs4 import BeautifulSoup
# imports excel library
from openpyxl import Workbook
# search results web page
url = ('https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&searchrcid'
       '=0305&searchcampusid=305&yrtr=20205&subject=ITEC&courseNumber=&courseId=&openValue=OPEN_PLUS_WA'
       'ITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=&credittype=ALL&credits=&instr'
       'uctor=&keyword=&begindate=&site=&resultNumber=250')
# gets response from web page
response = requests.get(url)
# uses the html.parser to parse the web page
soup = BeautifulSoup(response.content, 'html.parser')
# finds table header in html of the search results web page and stores in course_table variable
course_table = soup.find_all('table', id='resultsTable')
# first thing in the course_table variable
course_table = course_table[0]

# opens workbook
workbook = Workbook()
# first worksheet in the workbook
worksheet = workbook.active
# titles worksheet
worksheet.title = 'ITEC Courses'

# adds headers to spreadsheet columns
worksheet.cell(1, 1, 'ID #')
worksheet.cell(1, 2, 'Course Number')
worksheet.cell(1, 3, 'Section Number')
worksheet.cell(1, 4, 'Course Title')
worksheet.cell(1, 5, 'Day Class Meets')
worksheet.cell(1, 6, 'Time Class Meets')
worksheet.cell(1, 7, 'Credits')
worksheet.cell(1, 8, 'Instructor')

# sets spreadsheet rows = to 2 to skip row with headers
ss_row = 2
# finds data with the tag tr
for row in course_table.find_all('tr'):
    # creates empty list for data to be added later
    data = []
    # finds data with tag td and adds to empty list
    for cell in row.find_all('td'):
        data.append(cell.text.strip())
    # deletes all empty items from data
    for result in data:
        if result == '':
            data.remove(result)
    # adds data to spreadsheet
    if len(data):
        worksheet.cell(ss_row, 1, data[0])
        worksheet.cell(ss_row, 2, data[2])
        worksheet.cell(ss_row, 3, data[3])
        worksheet.cell(ss_row, 4, data[4])
        worksheet.cell(ss_row, 5, data[6])
        worksheet.cell(ss_row, 6, data[7])
        worksheet.cell(ss_row, 7, data[8])
        worksheet.cell(ss_row, 8, data[10])
        ss_row += 1
# saves spreadsheet
workbook.save('itec_courses.xlsx')

# Thanks again for the help and a great semester!
